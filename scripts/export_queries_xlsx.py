"""把 query 用例与产物导出为 xlsx 表格，便于人工审阅。

列布局：
    A: query id
    B: query 内容
    C: DSL 文件名（dsl/{id}.jsonl，缺失则标注）
    D: DSL 内容（全文，超过 Excel 单元格上限时截断并标注）
    E: output 内的 PNG 图片（嵌入单元格，缺失则标注）

用法（从仓库根目录运行）：

    python scripts/export_queries_xlsx.py --queries queries_100.jsonl
    python scripts/export_queries_xlsx.py --queries queries_100.jsonl --output export.xlsx
    python scripts/export_queries_xlsx.py --queries queries_100.jsonl --dsl-dir dsl --images output

依赖：openpyxl、Pillow（见 requirements 或本机自带）。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "Automation"))

from automation.config import safe_path_name  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.drawing.image import Image as XlImage  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from PIL import Image as PilImage  # noqa: E402

# Excel 单元格字符上限 32767，留余量
CELL_TEXT_LIMIT = 32000
# 嵌入图片显示的最大宽高（像素）
IMAGE_MAX_WIDTH = 240
IMAGE_MAX_HEIGHT = 320
# 列宽（字符数近似值）
COL_WIDTHS = {"A": 32, "B": 55, "C": 40, "D": 80, "E": 36}
HEADERS = ["query_id", "query", "dsl_file", "dsl_content", "image"]


def load_queries(path: Path) -> list[dict]:
    queries: list[dict] = []
    with path.open(encoding="utf-8-sig") as fh:
        for lineno, line in enumerate(fh, 1):
            line = line.strip()
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError as exc:
                raise SystemExit(f"{path}:{lineno} 不是合法 JSON：{exc}")
            if "id" not in item or "query" not in item:
                raise SystemExit(f"{path}:{lineno} 缺少 id 或 query 字段：{line[:80]}")
            queries.append(item)
    return queries


def find_image(images_dir: Path, stem: str) -> Path | None:
    for suffix in (".png", ".jpg", ".jpeg", ".webp"):
        candidate = images_dir / f"{stem}{suffix}"
        if candidate.is_file():
            return candidate
    return None


def scaled_image_size(path: Path) -> tuple[int, int]:
    """等比缩放到展示上限内，返回 (宽, 高) 像素。"""
    with PilImage.open(path) as img:
        width, height = img.size
    scale = min(IMAGE_MAX_WIDTH / width, IMAGE_MAX_HEIGHT / height, 1.0)
    return max(1, round(width * scale)), max(1, round(height * scale))


def main() -> int:
    parser = argparse.ArgumentParser(description="导出 query/DSL/图片到 xlsx")
    parser.add_argument("--queries", type=Path, required=True, help="queries JSONL 文件")
    parser.add_argument("--dsl-dir", type=Path, default=Path("dsl"), help="DSL 目录，默认 dsl/")
    parser.add_argument("--images", type=Path, default=Path("output"), help="图片目录，默认 output/")
    parser.add_argument("--output", type=Path, default=None, help="输出 xlsx，默认 <queries文件名>_export.xlsx")
    args = parser.parse_args()

    queries = load_queries(args.queries)
    output = args.output or args.queries.with_name(f"{args.queries.stem}_export.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "queries"

    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    for col, title in zip("ABCDE", HEADERS):
        cell = ws[f"{col}1"]
        cell.value = title
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[col].width = COL_WIDTHS[col]
    ws.freeze_panes = "A2"

    missing_dsl = 0
    missing_img = 0
    for row_index, item in enumerate(queries, start=2):
        qid = str(item["id"])
        stem = safe_path_name(qid)
        dsl_path = args.dsl_dir / f"{stem}.jsonl"
        image_path = find_image(args.images, stem)

        ws.cell(row=row_index, column=1, value=qid).alignment = wrap_top
        ws.cell(row=row_index, column=2, value=str(item["query"])).alignment = wrap_top

        if dsl_path.is_file():
            ws.cell(row=row_index, column=3, value=dsl_path.name).alignment = wrap_top
            content = dsl_path.read_text(encoding="utf-8-sig")
            if len(content) > CELL_TEXT_LIMIT:
                content = content[:CELL_TEXT_LIMIT] + f"\n... [内容过长已截断，完整见 {dsl_path.name}]"
            ws.cell(row=row_index, column=4, value=content).alignment = wrap_top
        else:
            missing_dsl += 1
            ws.cell(row=row_index, column=3, value="(缺失)").alignment = wrap_top

        if image_path:
            try:
                width, height = scaled_image_size(image_path)
                xl_image = XlImage(str(image_path))
                xl_image.width = width
                xl_image.height = height
                ws.add_image(xl_image, f"E{row_index}")
                # 行高按图片高度换算（1 像素 ≈ 0.75 磅），留少量边距
                ws.row_dimensions[row_index].height = height * 0.75 + 6
            except Exception as exc:
                ws.cell(row=row_index, column=5, value=f"(图片嵌入失败: {exc})").alignment = wrap_top
        else:
            missing_img += 1
            ws.cell(row=row_index, column=5, value="(缺失)").alignment = center

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"queries 总数 : {len(queries)}")
    print(f"缺 DSL       : {missing_dsl}")
    print(f"缺图片       : {missing_img}")
    print(f"输出文件     : {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
